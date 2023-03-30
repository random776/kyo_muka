import requests
from bs4 import BeautifulSoup

def receive():
    try:
        r = requests.get("https://www.c.u-tokyo.ac.jp/zenki/index.html")
        new_elem = set()
        soup = BeautifulSoup(r.text, "html.parser")
        try:
            import openpyxl # エクセルファイルを依代とする。
            old_elem_ = openpyxl.load_workbook('./old_elem.xlsx')
            sheet = old_elem_.active
            old_elem = set()
            for i in sheet.iter_rows(values_only="True"):
                old_elem.add(i)
            
            r = requests.get("https://www.c.u-tokyo.ac.jp/zenki/index.html")
            new_elem = set()
            new_elem2 = set()
            soup = BeautifulSoup(r.text, "html.parser")
            for dd in soup.find_all("dd"):
                if x := dd.find("a"):
                    new_elem2.add(x.text) # new_elemに要素を追加していく。
                    new_elem.add((x.text,)) # 比較用。どうしても() がついてしまうため。
            
            new_elem_ = openpyxl.load_workbook('./old_elem.xlsx')
            sheet = new_elem_.active
            for i, item in enumerate(new_elem2, 1):
                sheet[f"A{i}"] = item # 新しい鋳型に書き換える
                new_elem_.save('./old_elem.xlsx')

        except:
            pass
        
        if added := new_elem-old_elem:
            print("追加 :", added)
            a = added
            headers = {
            'Authorization': 'Bearer ラインnotify api のトークン',
            }

            files = {
            'message': (None, '{}   確認: https://www.c.u-tokyo.ac.jp/zenki/index.html (東京大学教養学部）'.format(a)),
            }

            requests.post('https://notify-api.line.me/api/notify', headers=headers, files=files)

        
        elif removed := old_elem-new_elem:
            print("削除 :", removed)


        else:
            print("教養学部HP:変化なし")
    
    except ConnectionError:
        pass
